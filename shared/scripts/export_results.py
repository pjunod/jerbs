"""
export_results.py — Job Email Screener
Converts screener results JSON into a formatted .xlsx / Google Sheets-ready file.

Usage:
    python export_results.py <results_json_file> <output_xlsx_file>

Or import and call:
    from export_results import export_to_xlsx
    export_to_xlsx(results, "job_screener_results.xlsx")
"""

import json
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Status pipeline ───────────────────────────────────────────────────────────

STATUS_PIPELINE = [
    "New",
    "Reply drafted",
    "Awaiting info",
    "Info received",
    "Interested — moving forward",
    "Intro / recruiter call scheduled",
    "Intro / recruiter call done",
    "Interview scheduled",
    "Interviewing",
    "Final round",
    "Offer received",
    "Negotiating",
    "Offer accepted",
    "Offer declined",
    "No response",
    "Withdrew",
    "Rejected",
    "Filtered out",
]

# Statuses that go into the collapsed "dead ends" group at the bottom
DEAD_END_STATUSES = {"Offer declined", "No response", "Withdrew", "Rejected", "Filtered out"}

STATUS_COLORS = {
    "New": ("EFF6FF", "1E40AF"),
    "Reply drafted": ("DBEAFE", "1D4ED8"),
    "Awaiting info": ("FEF9C3", "854D0E"),
    "Info received": ("FEF3C7", "92400E"),
    "Interested — moving forward": ("D1FAE5", "065F46"),
    "Intro / recruiter call scheduled": ("D1FAE5", "065F46"),
    "Intro / recruiter call done": ("A7F3D0", "064E3B"),
    "Interview scheduled": ("A7F3D0", "064E3B"),
    "Interviewing": ("6EE7B7", "064E3B"),
    "Final round": ("34D399", "022C22"),
    "Offer received": ("BBF7D0", "14532D"),
    "Negotiating": ("86EFAC", "14532D"),
    "Offer accepted": ("4ADE80", "14532D"),
    "Offer declined": ("F3F4F6", "6B7280"),
    "No response": ("F3F4F6", "6B7280"),
    "Withdrew": ("F3F4F6", "6B7280"),
    "Rejected": ("FEE2E2", "991B1B"),
    "Filtered out": ("F3F4F6", "9CA3AF"),
}

STATUS_GUIDE = {
    "New": "Just screened — no action taken yet",
    "Reply drafted": "Draft reply generated — ready to send",
    "Awaiting info": "Sent request for missing info — waiting on response",
    "Info received": "They replied — review and decide next step",
    "Interested — moving forward": "Decided to pursue — next step TBD",
    "Intro / recruiter call scheduled": "Call booked",
    "Intro / recruiter call done": "Call completed — assessing fit",
    "Interview scheduled": "Technical / hiring manager interview booked",
    "Interviewing": "In the interview process",
    "Final round": "Final interviews in progress",
    "Offer received": "Written offer in hand — reviewing",
    "Negotiating": "Actively negotiating comp / terms",
    "Offer accepted": "Signed",
    "Offer declined": "Passed on the offer",
    "No response": "Sent reply but heard nothing back",
    "Withdrew": "Removed yourself from consideration",
    "Rejected": "Rejected by the company",
    "Filtered out": "Did not meet screening criteria — no action",
}


def default_status(verdict):
    return "Filtered out" if verdict == "fail" else "New"


def is_dead_end(item):
    status = item.get("status") or default_status(item.get("verdict", "fail"))
    return status in DEAD_END_STATUSES


# ── Verdict display ───────────────────────────────────────────────────────────

VERDICT_COLORS = {
    "pass": {"bg": "D4EDDA", "fg": "155724"},
    "maybe": {"bg": "FFF3CD", "fg": "856404"},
    "fail": {"bg": "F8D7DA", "fg": "721C24"},
}
VERDICT_LABELS = {"pass": "Interested", "maybe": "Maybe", "fail": "Filtered out"}

# ── Column definitions ────────────────────────────────────────────────────────

COLUMNS = [
    ("Date screened", 16),
    ("Source", 18),
    ("Company", 22),
    ("Role", 32),
    ("Location", 18),
    ("Email date", 13),
    ("From", 26),
    ("Verdict", 13),
    ("Status", 26),
    ("Reason", 40),
    ("Dealbreaker", 26),
    ("Comp assessment", 40),
    ("Missing info", 32),
    ("Notes", 35),
    ("Draft reply", 55),
]

STATUS_COL = 9
NOTES_COL = 14
HEADER_BG = "1F2937"
HEADER_FG = "FFFFFF"
ALT_BG = "F9FAFB"
GROUP_BG = "E5E7EB"  # separator row background


def make_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def make_separator_border():
    thick = Side(style="medium", color="9CA3AF")
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thick, bottom=thick)


# ── Write a data row ──────────────────────────────────────────────────────────


def write_data_row(ws, row_idx, item, run_date, is_alt=False):
    verdict = item.get("verdict", "fail")
    status = item.get("status") or default_status(verdict)
    vc = VERDICT_COLORS.get(verdict, VERDICT_COLORS["fail"])
    sc = STATUS_COLORS.get(status, ("F3F4F6", "6B7280"))

    vals = [
        run_date,
        item.get("source", ""),
        item.get("company", ""),
        item.get("role", ""),
        item.get("location", ""),
        item.get("email_date", ""),
        item.get("from", ""),
        VERDICT_LABELS.get(verdict, verdict.title()),
        status,
        item.get("reason", ""),
        item.get("dealbreaker") or "",
        item.get("comp_assessment") or "",
        ", ".join(item.get("missing_fields") or []),
        item.get("notes") or "",
        item.get("reply_draft") or "",
    ]

    for ci, value in enumerate(vals, 1):
        cell = ws.cell(row=row_idx, column=ci, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = make_border()
        cell.font = Font(name="Arial", size=10)

        if ci == 8:  # Verdict
            cell.fill = PatternFill("solid", start_color=vc["bg"])
            cell.font = Font(bold=True, color=vc["fg"], name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="top")
        elif ci == STATUS_COL:  # Status
            bg, fg = sc
            cell.fill = PatternFill("solid", start_color=bg)
            cell.font = Font(bold=True, color=fg, name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        elif ci == NOTES_COL:  # Notes
            cell.fill = PatternFill("solid", start_color="FFFBEB")
        elif is_alt:
            cell.fill = PatternFill("solid", start_color=ALT_BG)

    ws.row_dimensions[row_idx].height = 60


def write_separator_row(ws, row_idx, label, count):
    """Write a collapsible group header row."""
    cell = ws.cell(row=row_idx, column=1, value=f"▸  {label}  ({count})")
    cell.font = Font(bold=True, color="4B5563", name="Arial", size=10, italic=True)
    cell.fill = PatternFill("solid", start_color=GROUP_BG)
    cell.alignment = Alignment(vertical="center")
    cell.border = make_separator_border()

    for ci in range(2, len(COLUMNS) + 1):
        c = ws.cell(row=row_idx, column=ci)
        c.fill = PatternFill("solid", start_color=GROUP_BG)
        c.border = make_separator_border()

    ws.row_dimensions[row_idx].height = 20


# ── Main export ───────────────────────────────────────────────────────────────


def export_to_xlsx(results_data, output_path):
    wb = Workbook()
    run_date = results_data.get("run_date", datetime.today().strftime("%Y-%m-%d"))
    items = results_data.get("results", [])
    counts = {"pass": 0, "maybe": 0, "fail": 0}
    for r in items:
        counts[r.get("verdict", "fail")] += 1

    # ── Summary sheet ─────────────────────────────────────────────────────────
    summary = wb.active
    summary.title = "Summary"

    rows = [
        ["Job Email Screener — Pipeline Tracker"],
        ["Run date", run_date],
        ["Total screened", len(items)],
        ["Interested", counts["pass"]],
        ["Maybe", counts["maybe"]],
        ["Filtered out", counts["fail"]],
        [],
        ["Status", "Meaning"],
    ] + [[s, STATUS_GUIDE[s]] for s in STATUS_PIPELINE]

    for ri, row in enumerate(rows, 1):
        for ci, val in enumerate(row, 1):
            cell = summary.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font = Font(bold=True, size=14, color=HEADER_BG)
            elif ri == 8:
                cell.font = Font(bold=True, color=HEADER_FG)
                cell.fill = PatternFill("solid", start_color=HEADER_BG)
            elif ri in (4, 5, 6) and ci == 2:
                vk = {4: "pass", 5: "maybe", 6: "fail"}[ri]
                c = VERDICT_COLORS[vk]
                cell.fill = PatternFill("solid", start_color=c["bg"])
                cell.font = Font(bold=True, color=c["fg"])
            elif ri > 8 and ci == 1:
                if val in STATUS_COLORS:
                    bg, fg = STATUS_COLORS[val]
                    cell.fill = PatternFill("solid", start_color=bg)
                    cell.font = Font(bold=True, color=fg, size=10)
            elif ci == 1:
                cell.font = Font(bold=True, color="374151")

    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 48

    # ── Results sheet ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Results")

    # Configure outline so collapse button appears above the group rows
    ws.sheet_properties.outlinePr.summaryBelow = False

    for ci, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = Font(bold=True, color=HEADER_FG, name="Arial")
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = make_border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Status dropdown
    opts = '","'.join(STATUS_PIPELINE)
    dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=False, showDropDown=False)
    dv.sqref = f"{get_column_letter(STATUS_COL)}2:{get_column_letter(STATUS_COL)}10000"
    ws.add_data_validation(dv)

    # Split items: active (visible) vs dead-end (collapsible)
    verdict_order = {"pass": 0, "maybe": 1, "fail": 2}
    all_items = sorted(items, key=lambda r: verdict_order.get(r.get("verdict", "fail"), 3))

    active_items = [r for r in all_items if not is_dead_end(r)]
    dead_end_items = [r for r in all_items if is_dead_end(r)]

    # Group dead-end items by status for sub-grouping
    dead_end_groups = {}
    for r in dead_end_items:
        status = r.get("status") or default_status(r.get("verdict", "fail"))
        dead_end_groups.setdefault(status, []).append(r)

    # Ordered dead-end group labels (only include ones that have items)
    dead_end_order = ["Offer declined", "No response", "Withdrew", "Rejected", "Filtered out"]
    ordered_dead_end_groups = [
        (s, dead_end_groups[s]) for s in dead_end_order if s in dead_end_groups
    ]

    # Write active rows
    current_row = 2
    for idx, item in enumerate(active_items):
        write_data_row(ws, current_row, item, run_date, is_alt=(idx % 2 == 0))
        current_row += 1

    # Write dead-end groups — each as a collapsible section
    if dead_end_items:
        # Leave one blank row as visual separator before groups
        ws.row_dimensions[current_row].height = 8
        current_row += 1

        for group_status, group_items in ordered_dead_end_groups:
            # Separator / group header row (NOT grouped — always visible, acts as the toggle label)
            write_separator_row(ws, current_row, group_status, len(group_items))
            current_row += 1

            # Data rows in this group — outlineLevel=1 makes them collapsible
            for idx, item in enumerate(group_items):
                write_data_row(ws, current_row, item, run_date, is_alt=(idx % 2 == 0))
                ws.row_dimensions[current_row].outlineLevel = 1
                ws.row_dimensions[current_row].hidden = True  # collapsed by default
                current_row += 1

    wb.save(output_path)
    print(
        f"Exported {len(items)} results → {len(active_items)} active, {len(dead_end_items)} collapsed | {output_path}"
    )


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python export_results.py <results.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1]) as f:
        data = json.load(f)
    export_to_xlsx(data, sys.argv[2])
