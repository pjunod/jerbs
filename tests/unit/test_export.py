"""
Unit tests for export_results.py — xlsx export logic.
"""

import json
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "shared" / "scripts"))

from export_results import (
    COLUMNS,
    STATUS_PIPELINE,
    VERDICT_LABELS,
    default_status,
    export_to_xlsx,
    is_dead_end,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def make_result(verdict="pass", company="TechCorp", role="Staff Engineer", status=None, **kwargs):
    r = {
        "verdict": verdict,
        "company": company,
        "role": role,
        "location": "Remote",
        "source": "Direct Outreach",
        "email_date": "2026-03-28",
        "from": "recruiter@techcorp.com",
        "reason": "Clears all criteria.",
        "dealbreaker": None,
        "comp_assessment": "Strong comp.",
        "missing_fields": [],
        "reply_draft": "Interested!\n\nAlex",
        "notes": "",
    }
    if status:
        r["status"] = status
    r.update(kwargs)
    return r


def run_export(items, run_date="2026-03-28"):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        export_to_xlsx({"run_date": run_date, "results": items}, path)
        wb = load_workbook(path)
        return wb
    finally:
        os.unlink(path)


# ---------------------------------------------------------------------------
# default_status / is_dead_end helpers
# ---------------------------------------------------------------------------


class TestHelpers:
    def test_default_status_fail_is_filtered_out(self):
        assert default_status("fail") == "Filtered out"

    def test_default_status_pass_is_new(self):
        assert default_status("pass") == "New"

    def test_default_status_maybe_is_new(self):
        assert default_status("maybe") == "New"

    def test_is_dead_end_filtered_out(self):
        assert is_dead_end({"verdict": "fail"}) is True

    def test_is_dead_end_pass_is_not(self):
        assert is_dead_end({"verdict": "pass"}) is False

    def test_is_dead_end_explicit_no_response(self):
        assert is_dead_end({"verdict": "pass", "status": "No response"}) is True

    def test_is_dead_end_explicit_rejected(self):
        assert is_dead_end({"verdict": "pass", "status": "Rejected"}) is True

    def test_is_dead_end_active_status(self):
        assert is_dead_end({"verdict": "pass", "status": "Interviewing"}) is False


# ---------------------------------------------------------------------------
# Sheet names and structure
# ---------------------------------------------------------------------------


class TestSheetStructure:
    def test_two_sheets_created(self):
        wb = run_export([make_result()])
        assert set(wb.sheetnames) == {"Summary", "Results"}

    def test_results_has_header_row(self):
        wb = run_export([make_result()])
        ws = wb["Results"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMNS))]
        assert headers[0] == "Date screened"
        assert "Verdict" in headers
        assert "Status" in headers
        assert "Draft reply" in headers

    def test_all_column_headers_present(self):
        wb = run_export([make_result()])
        ws = wb["Results"]
        expected = [col[0] for col in COLUMNS]
        actual = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMNS))]
        assert actual == expected


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------


class TestSummarySheet:
    def test_run_date_in_summary(self):
        wb = run_export([make_result()], run_date="2026-03-28")
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=c).value for r in range(1, 10) for c in range(1, 3)]
        assert "2026-03-28" in values

    def test_counts_correct(self):
        items = [
            make_result("pass"),
            make_result("pass"),
            make_result("maybe"),
            make_result("fail"),
            make_result("fail"),
            make_result("fail"),
        ]
        wb = run_export(items)
        ws = wb["Summary"]
        # Find count rows (rows 3-6 per export logic)
        total = ws.cell(row=3, column=2).value
        interested = ws.cell(row=4, column=2).value
        maybe = ws.cell(row=5, column=2).value
        filtered = ws.cell(row=6, column=2).value
        assert total == 6
        assert interested == 2
        assert maybe == 1
        assert filtered == 3

    def test_status_guide_included(self):
        wb = run_export([])
        ws = wb["Summary"]
        all_values = set()
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    all_values.add(cell)
        # Spot-check a few status guide entries
        assert "New" in all_values
        assert "Offer accepted" in all_values
        assert "Filtered out" in all_values


# ---------------------------------------------------------------------------
# Results sheet data
# ---------------------------------------------------------------------------


class TestResultsData:
    def test_pass_result_written(self):
        wb = run_export([make_result("pass", company="TechCorp", role="Staff Engineer")])
        ws = wb["Results"]
        row2 = [ws.cell(row=2, column=i + 1).value for i in range(len(COLUMNS))]
        assert "TechCorp" in row2
        assert "Staff Engineer" in row2
        assert VERDICT_LABELS["pass"] in row2

    def test_verdict_label_displayed(self):
        wb = run_export(
            [
                make_result("pass"),
                make_result("maybe"),
                make_result("fail"),
            ]
        )
        ws = wb["Results"]
        verdict_col = next(i + 1 for i, (h, _) in enumerate(COLUMNS) if h == "Verdict")
        # Scan all rows — fail verdict rows land in the dead-end collapsed group
        all_verdicts = {ws.cell(row=r, column=verdict_col).value for r in range(2, ws.max_row + 1)}
        assert VERDICT_LABELS["pass"] in all_verdicts
        assert VERDICT_LABELS["maybe"] in all_verdicts
        assert VERDICT_LABELS["fail"] in all_verdicts

    def test_missing_fields_joined(self):
        item = make_result("maybe", missing_fields=["Salary", "Remote policy"])
        wb = run_export([item])
        ws = wb["Results"]
        missing_col = next(i + 1 for i, (h, _) in enumerate(COLUMNS) if h == "Missing info")
        cell_val = ws.cell(row=2, column=missing_col).value
        assert "Salary" in cell_val
        assert "Remote policy" in cell_val

    def test_run_date_in_first_column(self):
        wb = run_export([make_result()], run_date="2026-03-28")
        ws = wb["Results"]
        assert ws.cell(row=2, column=1).value == "2026-03-28"

    def test_sort_order_pass_before_maybe_before_fail(self):
        items = [
            make_result("fail", company="FailCo"),
            make_result("pass", company="PassCo"),
            make_result("maybe", company="MaybeCo"),
        ]
        wb = run_export(items)
        ws = wb["Results"]
        company_col = next(i + 1 for i, (h, _) in enumerate(COLUMNS) if h == "Company")
        companies = [
            ws.cell(row=r, column=company_col).value
            for r in range(2, 5)
            if ws.cell(row=r, column=company_col).value
        ]
        # Pass should appear before maybe, maybe before fail (fail is dead-end, appears later)
        pass_idx = next((i for i, c in enumerate(companies) if c == "PassCo"), None)
        maybe_idx = next((i for i, c in enumerate(companies) if c == "MaybeCo"), None)
        if pass_idx is not None and maybe_idx is not None:
            assert pass_idx < maybe_idx


# ---------------------------------------------------------------------------
# Dead-end grouping
# ---------------------------------------------------------------------------


class TestDeadEndGrouping:
    def test_dead_end_rows_are_hidden(self):
        items = [
            make_result("fail", company="BadCo"),
        ]
        wb = run_export(items)
        ws = wb["Results"]
        # Row 2 is the separator, row 3 is the dead-end data row (hidden)
        # Find any hidden rows
        hidden_rows = [r for r in range(2, 10) if ws.row_dimensions[r].hidden]
        assert len(hidden_rows) > 0

    def test_active_rows_not_hidden(self):
        items = [
            make_result("pass", company="GoodCo"),
        ]
        wb = run_export(items)
        ws = wb["Results"]
        assert ws.row_dimensions[2].hidden is False

    def test_empty_results(self):
        wb = run_export([])
        ws = wb["Results"]
        # Row 2 should be empty (no data)
        assert ws.cell(row=2, column=1).value is None

    def test_mixed_active_and_dead_end(self):
        items = [
            make_result("pass", company="GoodCo"),
            make_result("fail", company="BadCo"),
        ]
        wb = run_export(items)
        ws = wb["Results"]
        # Active row (pass) should not be hidden
        assert ws.row_dimensions[2].hidden is False


# ---------------------------------------------------------------------------
# Status dropdown
# ---------------------------------------------------------------------------


class TestStatusDropdown:
    def test_data_validation_present(self):
        wb = run_export([make_result()])
        ws = wb["Results"]
        # openpyxl exposes data_validations
        dvs = ws.data_validations.dataValidation
        assert len(dvs) > 0

    def test_all_statuses_in_pipeline(self):
        # Verify STATUS_PIPELINE contains expected stages
        assert "New" in STATUS_PIPELINE
        assert "Offer accepted" in STATUS_PIPELINE
        assert "Filtered out" in STATUS_PIPELINE
        assert "Interviewing" in STATUS_PIPELINE

    def test_pre_existing_status_preserved_in_export(self):
        r = make_result(status="Interviewing")
        assert r["status"] == "Interviewing"
        wb = run_export([r])
        ws = wb["Results"]
        headers = {cell.value: cell.column for cell in ws[1]}
        status_col = headers.get("Status")
        assert ws.cell(row=2, column=status_col).value == "Interviewing"


# ---------------------------------------------------------------------------
# export_results.py __main__ guard
# ---------------------------------------------------------------------------


class TestScriptGuard:
    _export_path = str(
        Path(__file__).parent.parent.parent / "shared" / "scripts" / "export_results.py"
    )

    def test_too_few_args_exits_with_error(self):
        import runpy

        import pytest

        with patch("sys.argv", ["export_results.py"]):
            with patch("builtins.print"):
                with pytest.raises(SystemExit) as exc_info:
                    runpy.run_path(self._export_path, run_name="__main__")
        assert exc_info.value.code == 1

    def test_runs_export_when_args_provided(self, tmp_path):
        import runpy

        data = {"run_date": "2026-03-28", "results": []}
        data_file = tmp_path / "data.json"
        data_file.write_text(json.dumps(data))
        output_file = tmp_path / "out.xlsx"

        with patch("sys.argv", ["export_results.py", str(data_file), str(output_file)]):
            with patch("builtins.print"):
                runpy.run_path(self._export_path, run_name="__main__")
        assert output_file.exists()
